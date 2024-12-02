import openpyxl
import pandas as pd

def excel_to_dataframe_with_formulas(file_path, sheets=None, columns=None):
    """
    Convierte un archivo de Excel a un DataFrame, incluyendo fórmulas o valores según corresponda.
    Incluye una columna indicando la hoja, fila y columna donde se encuentra cada dato.

    Args:
        file_path (str): Ruta del archivo Excel.
        sheets (list): Lista de hojas a incluir, por ejemplo ['Hoja1', 'Hoja2']. Si es None, incluye todas las hojas.
        columns (list): Lista de columnas a incluir, por ejemplo ['A', 'C']. Si es None, incluye todas las columnas.

    Returns:
        pandas.DataFrame: Contenido del archivo en formato DataFrame.
    """
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    data = []

    # Determinar las hojas a procesar
    sheetnames_to_process = sheets if sheets else workbook.sheetnames

    for sheet in sheetnames_to_process:
        if sheet not in workbook.sheetnames:
            print(f"Advertencia: La hoja '{sheet}' no existe en el archivo.")
            continue

        worksheet = workbook[sheet]

        for row in worksheet.iter_rows():
            for cell in row:
                # Manejar el caso de celdas combinadas
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue  # Omitir celdas combinadas

                # Incluir solo las columnas especificadas (o todas si columns es None)
                if columns is None or cell.column_letter in columns:
                    cell_data = {
                        "Sheet": sheet,
                        "Cell": cell.coordinate,
                        "Row": cell.row,
                        "Column": cell.column_letter,
                        "Value": cell.value,
                        "Formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                    }
                    data.append(cell_data)

    # Crear el DataFrame
    df = pd.DataFrame(data)

    # Eliminar filas donde Value es None
    df = df[df["Value"].notna()]

    # Reorganizar las columnas para mayor claridad
    df = df[["Sheet", "Cell", "Row", "Column", "Value", "Formula"]]

    return df


# Uso
file_path = "../Listado arranque Sabón con señales DCS v2.xlsx"  # Reemplaza con tu archivo
sheets_to_include = ["Lista confirmación arranque"]  # Especifica las hojas a procesar
columns_to_include = ["E", "F", "G"]  # Especifica las columnas a incluir

df = excel_to_dataframe_with_formulas(file_path, sheets_to_include, columns_to_include)

print("Conversión completada.")

